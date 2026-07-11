"""Ingestion Pipeline — Excel → 归一化 → Postgres + Qdrant（P2a / MVS 核心）。

核心闭环（MVS 验收：「上传乱列名 Excel → 问答命中该数据」）：

    上传 Excel
      └─ 解析为行（header → value）
         └─ identity 字段映射归一化 → CanonicalDocument 列表
            ├─ 落库 Postgres：raw_documents + canonical_documents + document_chunks
            └─ 嵌入 + 写入 Qdrant（同一 collection，与 RAG 共用）

所有外部依赖（DB session / 向量库 / 嵌入模型）均以参数注入，便于测试用
内存实现替换，无需真实 Postgres / Qdrant / BGE-M3。
"""

from __future__ import annotations

import hashlib
import io
import json
import logging
from datetime import date, datetime
from pathlib import Path
from typing import Any

from sqlalchemy.ext.asyncio import AsyncSession

from agents.ingestion_agent.chunking import _estimate_tokens, build_text_chunks, render_table
from agents.ingestion_agent.database.models import (
    CanonicalDocument as CanonicalDocumentORM,
)
from agents.ingestion_agent.database.models import DocumentChunk, RawDocument
from agents.ingestion_agent.fts import index_canonical
from agents.ingestion_agent.mapping_loader import (
    build_identity_mapping,
    normalize_field_name,
    normalize_rows,
)
from agents.ingestion_agent.normalization import normalize_table_rows, normalize_text_block
from agents.ingestion_agent.parsers import BlockType, parse_file
from agents.ingestion_agent.storage import (
    ObjectStorage,
    compute_file_hash,
    make_storage_key,
)
from agents.rag_agent.embeddings import EmbeddingModel
from agents.rag_agent.vector_store import CollectionConfig, VectorRecord, VectorStore

logger = logging.getLogger("fde.ingestion.pipeline")

# 归一化文档类型默认值（MVS 零配置）。
DEFAULT_DOC_TYPE = "excel_upload"
DEFAULT_SOURCE_SYSTEM = "excel_upload"
# 入库向量集合名（与 RAG 共用 fde_documents）。
DEFAULT_COLLECTION = "fde_documents"


def _to_json_safe(value: Any) -> Any:
    """把 Excel 单元格值转换为 JSON 安全类型（datetime/date → ISO 字符串）。"""
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value)


def render_canonical_text(cd: Any) -> str:
    """把一个归一化文档渲染为可嵌入 / 可检索的文本块。"""
    lines = [cd.title]
    for key, value in cd.fields.items():
        safe = _to_json_safe(value)
        if safe is None or safe == "":
            continue
        lines.append(f"{key}: {safe}")
    return "\n".join(lines).strip()


def compute_content_hash(payload: dict[str, Any]) -> str:
    """对归一化 payload 计算稳定哈希，用于幂等 / 去重（P3b 将进一步使用）。"""
    canonical = json.dumps(payload, sort_keys=True, ensure_ascii=False, default=str)
    return hashlib.sha256(canonical.encode("utf-8")).hexdigest()[:64]


async def _find_existing_raw(session: AsyncSession, file_hash: str) -> Any | None:
    """按文件级 content_hash 查是否已入库（P3b 幂等：重复上传不产生幽灵文档）。"""
    from sqlalchemy import select

    from agents.ingestion_agent.database.models import RawDocument as RawDoc

    res = await session.execute(select(RawDoc).where(RawDoc.content_hash == file_hash))
    return res.scalars().first()


class IngestionPipeline:
    """Excel 入库流水线（无状态，依赖注入）。"""

    @staticmethod
    def _improve_title(cd: Any, index: int) -> str:
        """当归一化未提供可读 title 时，用首字段值或行号兜底。"""
        if cd.title and cd.title != cd.doc_type:
            return str(cd.title)
        for value in cd.fields.values():
            if value not in (None, ""):
                return str(value)
        return f"{cd.doc_type} #{index + 1}"

    @staticmethod
    async def ingest_rows(
        headers: list[str],
        rows: list[dict[str, Any]],
        *,
        doc_type: str = DEFAULT_DOC_TYPE,
        source_ref: str | None = None,
        session: AsyncSession,
        vector_store: VectorStore,
        embedding_model: EmbeddingModel,
        storage: ObjectStorage | None = None,
        file_hash: str | None = None,
        storage_ref: str | None = None,
    ) -> dict[str, Any]:
        """把一组 Excel 行归一化、落库 Postgres 并写入 Qdrant。

        Args:
            headers: 列名列表。
            rows: 每行一个 ``{header: value}`` 字典。
            doc_type: 归一化文档类型（默认 ``excel_upload``）。
            source_ref: 数据源引用（如 ``local://<filename>``）。
            session: 异步数据库会话。
            vector_store: Qdrant 向量库（需实现 ``async_create_collection`` / ``async_upsert``）。
            embedding_model: 嵌入模型（需实现 ``encode_documents`` / ``get_dimension``）。

        Returns:
            入库摘要：doc_type / row 数 / canonical 数 / raw_id。
        """
        if not rows:
            raise ValueError("没有可入库的数据行（Excel 仅含表头或为空）")

        # 0) 文件级幂等（P3b）：相同原始字节（file_hash）不重复产生 RawDocument / 幽灵文档。
        if file_hash is not None:
            existing = await _find_existing_raw(session, file_hash)
            if existing is not None:
                return {
                    "doc_type": doc_type,
                    "source_ref": source_ref,
                    "rows": len(rows),
                    "canonical": 0,
                    "indexed_vectors": 0,
                    "raw_id": existing.id,
                    "duplicated": True,
                    "storage_ref": existing.storage_ref,
                }

        source_ref = source_ref or f"local://{doc_type}"
        canonical_docs = normalize_rows(headers, rows, doc_type=doc_type)

        # 1) raw_documents：保留原始抽取（元数据；原始字节存对象存储 storage_ref）
        raw = RawDocument(
            source_type=DEFAULT_SOURCE_SYSTEM,
            source_ref=source_ref,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            content_hash=file_hash,
            storage_ref=storage_ref,
            raw_payload={"headers": headers, "row_count": len(rows)},
        )
        session.add(raw)
        await session.flush()

        # 2) canonical_documents + document_chunks + Qdrant 点
        points: list[VectorRecord] = []
        for idx, cd in enumerate(canonical_docs):
            title = IngestionPipeline._improve_title(cd, idx)
            safe_fields = {k: _to_json_safe(v) for k, v in cd.fields.items()}
            content_hash = compute_content_hash(safe_fields)

            orm = CanonicalDocumentORM(
                raw_document_id=raw.id,
                doc_type=cd.doc_type,
                title=title,
                canonical_payload=safe_fields,
                storage_ref=f"{source_ref}#{idx}",
                source_connector=DEFAULT_SOURCE_SYSTEM,
                content_hash=content_hash,
            )
            session.add(orm)
            await session.flush()
            await index_canonical(session, orm, raw_document_id=raw.id)

            text = render_canonical_text(type("CD", (), {"title": title, "fields": safe_fields})())
            session.add(
                DocumentChunk(
                    canonical_document_id=orm.id,
                    chunk_index=0,
                    content=text,
                    embedding_id=orm.id,
                    token_count=_estimate_tokens(text),
                )
            )

            emb = (await embedding_model.encode_documents([text]))[0]
            points.append(
                VectorRecord(
                    id=orm.id,
                    vector=emb,
                    payload={
                        "text": text,
                        "doc_type": cd.doc_type,
                        "title": title,
                        "canonical": safe_fields,
                        "source": source_ref,
                    },
                )
            )

        await session.commit()

        # 3) 确保 collection 存在并 upsert
        collection = DEFAULT_COLLECTION
        cfg = getattr(vector_store, "config", None)
        if cfg is not None and getattr(cfg, "collection_name", None):
            collection = cfg.collection_name
        await vector_store.async_create_collection(
            CollectionConfig(name=collection, vector_size=embedding_model.get_dimension())
        )
        upserted = await vector_store.async_upsert(points, collection=collection)

        return {
            "doc_type": doc_type,
            "source_ref": source_ref,
            "rows": len(rows),
            "canonical": len(canonical_docs),
            "indexed_vectors": upserted,
            "raw_id": raw.id,
        }

    @staticmethod
    async def ingest_excel(
        data: bytes,
        filename: str,
        *,
        doc_type: str | None = None,
        source_ref: str | None = None,
        session: AsyncSession,
        vector_store: VectorStore,
        embedding_model: EmbeddingModel,
        storage: ObjectStorage | None = None,
    ) -> dict[str, Any]:
        """从 Excel 字节解析为行，再调用 :meth:`ingest_rows`。

        Args:
            data: ``.xlsx`` 文件字节。
            filename: 原始文件名（用于推导默认 doc_type / source_ref）。
            doc_type: 可选覆盖默认文档类型。
            source_ref: 可选覆盖数据源引用。
            session / vector_store / embedding_model: 注入依赖。

        Returns:
            同 :meth:`ingest_rows`。
        """
        try:
            import openpyxl
        except ImportError as exc:  # pragma: no cover - openpyxl 为声明依赖
            raise RuntimeError(
                "openpyxl 未安装，无法解析 Excel。请执行 pip install openpyxl"
            ) from exc

        # 文件级幂等 + 原始字节外置（P3b）
        file_hash = compute_file_hash(data)
        existing = await _find_existing_raw(session, file_hash)
        if existing is not None:
            eff_doc = doc_type or Path(filename).stem or DEFAULT_DOC_TYPE
            eff_ref = source_ref or f"local://{filename}"
            return {
                "doc_type": eff_doc,
                "source_ref": eff_ref,
                "rows": 0,
                "canonical": 0,
                "indexed_vectors": 0,
                "raw_id": existing.id,
                "duplicated": True,
                "storage_ref": existing.storage_ref,
            }
        storage_ref: str | None = None
        if storage is not None:
            storage_ref = await storage.put(make_storage_key(filename, file_hash), data)

        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            raise ValueError("Excel 文件无有效工作表")

        iter_rows = ws.iter_rows(values_only=True)
        try:
            header_row = next(iter_rows)
        except StopIteration:
            raise ValueError("Excel 文件为空或无法读取表头")

        headers = [
            (str(h).strip() if h is not None and str(h).strip() else f"column_{i}")
            for i, h in enumerate(header_row)
        ]
        # 归一化为合法字段名，避免重复列名冲突（后写覆盖，MVS 可接受）
        seen: dict[str, int] = {}
        norm_headers = []
        for h in headers:
            base = normalize_field_name(h)
            if base in seen:
                seen[base] += 1
                base = f"{base}_{seen[base]}"
            else:
                seen[base] = 0
            norm_headers.append(base)

        rows: list[dict[str, Any]] = []
        for raw_row in iter_rows:
            if raw_row is None or all(v is None for v in raw_row):
                continue
            row = {
                norm_headers[i]: (raw_row[i] if i < len(raw_row) else None)
                for i in range(len(norm_headers))
            }
            rows.append(row)

        effective_doc_type = doc_type or (Path(filename).stem or DEFAULT_DOC_TYPE)
        effective_source_ref = source_ref or f"local://{filename}"
        return await IngestionPipeline.ingest_rows(
            norm_headers,
            rows,
            doc_type=effective_doc_type,
            source_ref=effective_source_ref,
            session=session,
            vector_store=vector_store,
            embedding_model=embedding_model,
            storage=storage,
            file_hash=file_hash,
            storage_ref=storage_ref,
        )

    @staticmethod
    async def ingest_file(
        filename: str,
        data: bytes,
        *,
        doc_type: str | None = None,
        source_ref: str | None = None,
        session: AsyncSession,
        vector_store: VectorStore,
        embedding_model: EmbeddingModel,
        storage: ObjectStorage | None = None,
    ) -> dict[str, Any]:
        """解析任意受支持的文件（pdf/docx/pptx/csv/xlsx…）→ 三层归一化 → 父子
        chunk → 落库 Postgres + 进 Qdrant（P2b / 完整本地文件入库）。

        Args:
            filename: 原始文件名（用于类型判定与默认 doc_type / source_ref）。
            data: 文件字节。
            doc_type: 可选覆盖默认文档类型（默认按文件类型推导）。
            source_ref: 可选覆盖数据源引用。
            session / vector_store / embedding_model: 注入依赖。

        Returns:
            入库摘要：doc_type / filename / blocks / canonical / chunks / raw_id 等。
        """
        parsed = parse_file(filename, data)
        effective_doc_type = doc_type or parsed.doc_type or Path(filename).stem or DEFAULT_DOC_TYPE
        effective_source_ref = source_ref or f"local://{filename}"

        # 文件级幂等 + 原始字节外置（P3b）
        file_hash = compute_file_hash(data)
        existing = await _find_existing_raw(session, file_hash)
        if existing is not None:
            return {
                "doc_type": effective_doc_type,
                "source_ref": effective_source_ref,
                "filename": filename,
                "blocks": len(parsed.blocks),
                "canonical": 0,
                "chunks": 0,
                "indexed_vectors": 0,
                "raw_id": existing.id,
                "duplicated": True,
                "storage_ref": existing.storage_ref,
            }
        storage_ref: str | None = None
        if storage is not None:
            storage_ref = await storage.put(make_storage_key(filename, file_hash), data)

        raw = RawDocument(
            source_type="file_upload",
            source_ref=effective_source_ref,
            content_type=parsed.meta.get("content_type"),
            content_hash=file_hash,
            storage_ref=storage_ref,
            raw_payload={
                "filename": filename,
                "blocks": len(parsed.blocks),
                "file_type": parsed.meta.get("file_type"),
            },
        )
        session.add(raw)
        await session.flush()

        # pending: (chunk_id, child_text, payload) —— 先收齐，提交后再批量嵌入。
        pending: list[tuple[str, str, dict[str, Any]]] = []
        canonical_count = 0

        for bi, block in enumerate(parsed.blocks):
            if block.kind == BlockType.TABLE:
                headers = block.table_headers or []
                rows = block.table or []
                cds = normalize_table_rows(headers, rows, doc_type=effective_doc_type)
                for ri, cd in enumerate(cds):
                    title = IngestionPipeline._improve_title(cd, canonical_count)
                    safe = {k: _to_json_safe(v) for k, v in cd.fields.items()}
                    content_hash = compute_content_hash(safe)
                    orm = CanonicalDocumentORM(
                        raw_document_id=raw.id,
                        doc_type=cd.doc_type,
                        title=title,
                        canonical_payload=safe,
                        storage_ref=f"{effective_source_ref}#t{bi}r{ri}",
                        source_connector="file_upload",
                        content_hash=content_hash,
                    )
                    session.add(orm)
                    await session.flush()
                    await index_canonical(session, orm, raw_document_id=raw.id)
                    canonical_count += 1

                    child_text = "\n".join(
                        f"{k}: {safe[k]}" for k in safe if safe[k] not in (None, "")
                    )
                    parent_text = render_table(headers, rows)
                    chunk = DocumentChunk(
                        canonical_document_id=orm.id,
                        chunk_index=0,
                        content=child_text,
                        token_count=_estimate_tokens(child_text),
                    )
                    session.add(chunk)
                    await session.flush()
                    chunk.embedding_id = chunk.id
                    pending.append(
                        (
                            chunk.id,
                            child_text,
                            {
                                "block_kind": "table",
                                "row_index": ri,
                                "doc_type": cd.doc_type,
                                "title": title,
                                "text": child_text,
                                "parent_text": parent_text,
                                "canonical": safe,
                                "source": effective_source_ref,
                                "raw_id": raw.id,
                                **block.loc,
                            },
                        )
                    )
            elif block.kind in (BlockType.TEXT, BlockType.HEADING):
                clean = normalize_text_block(block.text)
                if not clean.strip():
                    continue
                title = f"{Path(filename).name} {block.loc}" if block.loc else effective_source_ref
                content_hash = compute_content_hash({"text": clean})
                orm = CanonicalDocumentORM(
                    raw_document_id=raw.id,
                    doc_type=effective_doc_type,
                    title=title,
                    canonical_payload={"text": clean[:2000]},
                    storage_ref=f"{effective_source_ref}#b{bi}",
                    source_connector="file_upload",
                    content_hash=content_hash,
                )
                session.add(orm)
                await session.flush()
                await index_canonical(session, orm, raw_document_id=raw.id)
                canonical_count += 1

                child_specs = build_text_chunks(
                    clean,
                    doc_type=effective_doc_type,
                    source_ref=effective_source_ref,
                    raw_id=raw.id,
                    loc=block.loc,
                )
                for ci, spec in enumerate(child_specs):
                    chunk = DocumentChunk(
                        canonical_document_id=orm.id,
                        chunk_index=ci,
                        content=spec.child_text,
                        token_count=_estimate_tokens(spec.child_text),
                    )
                    session.add(chunk)
                    await session.flush()
                    chunk.embedding_id = chunk.id
                    pending.append(
                        (
                            chunk.id,
                            spec.child_text,
                            {
                                **spec.metadata,
                                "title": title,
                                "text": spec.child_text,
                                "parent_text": spec.parent_text,
                                "canonical": {"text": clean[:2000]},
                                "source": effective_source_ref,
                                "doc_type": effective_doc_type,
                            },
                        )
                    )

        await session.commit()

        child_texts = [p[1] for p in pending]
        points = []
        if child_texts:
            vecs = await embedding_model.encode_documents(child_texts)
            collection = DEFAULT_COLLECTION
            cfg = getattr(vector_store, "config", None)
            if cfg is not None and getattr(cfg, "collection_name", None):
                collection = cfg.collection_name
            for (chunk_id, _text, payload), vec in zip(pending, vecs, strict=False):
                points.append(VectorRecord(id=chunk_id, vector=vec, payload=payload))
            await vector_store.async_create_collection(
                CollectionConfig(name=collection, vector_size=embedding_model.get_dimension())
            )
            await vector_store.async_upsert(points, collection=collection)

        return {
            "doc_type": effective_doc_type,
            "source_ref": effective_source_ref,
            "filename": filename,
            "blocks": len(parsed.blocks),
            "canonical": canonical_count,
            "chunks": len(pending),
            "indexed_vectors": len(points),
            "raw_id": raw.id,
        }


__all__ = [
    "IngestionPipeline",
    "build_identity_mapping",
    "compute_content_hash",
    "normalize_field_name",
    "render_canonical_text",
]
